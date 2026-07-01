from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
import requests
import json
import base64
import io
import csv
from markupsafe import Markup

class AuditTest(models.Model):
    _name = 'audit.test'
    _description = 'Test Audit'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    name = fields.Char(string='Titre', required=True, tracking=True)
    date = fields.Date(string='Date', required=True, default=fields.Date.context_today, tracking=True)
    
    # Booleans for tests
    exactitude = fields.Boolean(string='Exactitude', tracking=True)
    cutoff = fields.Boolean(string='Cut-off', tracking=True)
    mention_legale = fields.Boolean(string='Mention Légale', tracking=True)
    
    # Files
    facture_ids = fields.Many2many(
        'ir.attachment',
        'audit_test_facture_rel',
        'test_id',
        'attachment_id',
        string='Factures PDF',
        domain="[('mimetype', 'in', ['application/pdf'])]"
    )
    
    compta_excel = fields.Binary(string='Sélections de comptabilité (Excel)', attachment=True)
    compta_excel_filename = fields.Char(string='Nom Compta Excel')

    def action_verify_exactitude(self):
        self.ensure_one()
        if not self.facture_ids or not self.compta_excel:
            raise ValidationError("Veuillez d'abord uploader au moins une facture (PDF) ET les sélections de comptabilité (Excel).")

        api_key = self.env['ir.config_parameter'].sudo().get_param('tresorerie_chq.gemini_key')
        if not api_key:
            raise ValidationError("La clé API Google Gemini n'est pas configurée (tresorerie_chq.gemini_key).")

        # Lecture du fichier Excel
        try:
            import openpyxl
            excel_bytes = base64.b64decode(self.compta_excel)
            wb = openpyxl.load_workbook(filename=io.BytesIO(excel_bytes), data_only=True)
            
            sheet_name = None
            for sn in wb.sheetnames:
                if 'sélection' in sn.lower() or 'selection' in sn.lower():
                    sheet_name = sn
                    break
            
            if not sheet_name:
                sheet = wb.active
            else:
                sheet = wb[sheet_name]
            
            csv_data = io.StringIO()
            writer = csv.writer(csv_data)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow([str(c) if c is not None else '' for c in row])
            
            compta_text = csv_data.getvalue()
        except Exception as e:
            raise ValidationError(f"Erreur lors de la lecture du fichier Excel : {str(e)}")

        prompt_text = f"""Vous êtes un auditeur comptable. Vous avez reçu :
1. Une facture en pièce jointe (PDF)
2. Les données de la feuille 'sélection' de la comptabilité au format CSV ci-dessous :

DONNÉES COMPTABLES (CSV) :
{compta_text[:20000]}

Votre tâche :
1. Lire le numéro de la facture et son montant total TTC (ou à payer) dans le PDF de la facture.
2. Chercher ce numéro de facture dans les données comptables CSV fournies ci-dessus. Sachez que le numéro de la facture se trouve spécifiquement dans la colonne 'Numéro Fact' (qui correspond à la colonne H dans le fichier d'origine).
3. Extraire le montant qui est associé à cette facture dans le CSV.
4. Vérifier si les deux montants sont égaux (attention aux formats de nombres, ex: 1000.50 et 1 000,50).
5. Lire la date de la facture sur le document PDF et l'extraire au format YYYY-MM-DD.
6. Vérifier si TOUTES les mentions légales suivantes sont présentes sur la facture PDF :
   - Dénomination sociale de l'acheteur et du vendeur
   - Montant du capital social
   - Siège social
   - N° RC (Registre du Commerce)
   - N° IF (Identifiant Fiscal)
   - N° TP (Taxe Professionnelle ou Patente)
   - N° ICE (Identifiant Commun de l'Entreprise)
   - N° CNSS
   - Mention TVA
   - Mode de paiement

Retournez UNIQUEMENT un objet JSON valide, sans markdown.
Exemple attendu :
{{
    "invoice_number": "le numéro trouvé sur la facture PDF",
    "invoice_amount": le_montant_trouvé_sur_la_facture_en_nombre,
    "compta_amount": le_montant_trouvé_dans_la_compta_en_nombre_ou_null,
    "is_exact": true_ou_false,
    "invoice_date": "YYYY-MM-DD",
    "has_legal_mentions": true_ou_false
}}"""

        all_exactitude = True
        all_cutoff = True
        all_mentions = True
        full_message = "<h3>Résultat de l'Audit</h3><br/>"

        for facture in self.facture_ids:
            full_message += f"<b>Facture: {facture.name}</b><ul>"
            
            # Upload PDF to Gemini
            pdf_bytes = base64.b64decode(facture.datas)
            upload_url = f"https://generativelanguage.googleapis.com/upload/v1beta/files?key={api_key}"
            upload_headers = {
                "X-Goog-Upload-Protocol": "raw",
                "X-Goog-Upload-Header-Content-Type": "application/pdf",
                "Content-Type": "application/pdf"
            }
            try:
                upload_resp = requests.post(upload_url, headers=upload_headers, data=pdf_bytes, timeout=120)
                if upload_resp.status_code != 200:
                    err_msg = upload_resp.json().get("error", {}).get("message", upload_resp.text)
                    raise ValidationError(f"Erreur d'upload pour {facture.name} : {err_msg}")
                file_uri = upload_resp.json().get("file", {}).get("uri")
            except Exception as e:
                raise ValidationError(f"Erreur upload {facture.name} : {str(e)}")

            # Gemini Prompt
            payload = {
                "contents": [
                    {
                        "parts": [
                            {"text": prompt_text},
                            {"fileData": {"mimeType": "application/pdf", "fileUri": file_uri}}
                        ]
                    }
                ],
                "generationConfig": {"responseMimeType": "application/json", "temperature": 0.0}
            }
            
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent?key={api_key}"
            try:
                resp = requests.post(url, headers={"Content-Type": "application/json"}, json=payload, timeout=120)
                if resp.status_code != 200:
                    err_msg = resp.json().get("error", {}).get("message", resp.text)
                    raise ValidationError(f"Erreur Gemini pour {facture.name} : {err_msg}")
                raw_content = resp.json()["candidates"][0]["content"]["parts"][0]["text"]
                result = json.loads(raw_content)
            except Exception as e:
                raise ValidationError(f"Erreur analyse Gemini pour {facture.name} : {str(e)}")

            # Evaluate Results
            # 1. Exactitude
            if result.get("is_exact") is True:
                full_message += f"<li><span style='color: green;'>Exactitude : Succès</span> (Montant {result.get('invoice_amount')} vérifié)</li>"
            else:
                all_exactitude = False
                full_message += f"<li><span style='color: red;'>Exactitude : Échec</span> (Trouvé {result.get('invoice_amount')} vs Compta {result.get('compta_amount')})</li>"

            # 2. Cut-off
            inv_date = result.get("invoice_date")
            if inv_date and self.date:
                try:
                    inv_year = int(inv_date.split('-')[0])
                    if inv_year == self.date.year:
                        full_message += f"<li><span style='color: green;'>Cut-off : Succès</span> (Année {inv_year})</li>"
                    else:
                        all_cutoff = False
                        full_message += f"<li><span style='color: red;'>Cut-off : Échec</span> (Facture {inv_year} vs Test {self.date.year})</li>"
                except:
                    all_cutoff = False
                    full_message += f"<li><span style='color: orange;'>Cut-off : Erreur de date</span> ({inv_date})</li>"
            else:
                all_cutoff = False
                full_message += "<li><span style='color: orange;'>Cut-off : Date introuvable</span></li>"

            # 3. Mentions Légales
            if result.get("has_legal_mentions") is True:
                full_message += "<li><span style='color: green;'>Mentions Légales : Succès</span></li>"
            else:
                all_mentions = False
                full_message += "<li><span style='color: red;'>Mentions Légales : Échec</span> (Manquantes)</li>"
            
            full_message += "</ul>"

        # Update record booleans
        self.exactitude = all_exactitude
        self.cutoff = all_cutoff
        self.mention_legale = all_mentions

        self.message_post(body=Markup(full_message))
